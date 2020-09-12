import tkinter.messagebox
import funcs
import pyproj
from openpyxl import *

datum = pyproj.Proj('+proj=tmerc +lat_0=39.66825833333333 +lon_0=-8.133108333333334 +k=1 +x_0=0 +y_0=0 +ellps=GRS80 +units=m +no_defs')#PT-TM06/ETRS89 EPSG:3763
wgs84 = pyproj.Proj('+init=EPSG:4326') # LatLon with WGS84 datum used by GPS units and Google Earth

######################################### BLOCOS NOS ######################################################
###########################################################################################################

def rede_pt():

    tkinter.messagebox.showinfo('Abertura de ficheiro', 'Seleccionar ficheiro txt com\ntabela de pontos extraídos do kml.')

    opcao = 'rede PT'
    #mensagem='\nEscolha: '+opcao
    #funcs.escrever(mensagem)

    funcs.t_fich="text files"
    funcs.e_fich="*.txt"

    funcs.abrir()
    if funcs.ficheiro == '':
        tkinter.messagebox.showinfo('Erro', 'A abertura do ficheiro falhou')
        return 0

    texto = ''

    str1 = 'A'
    str2 = 'EDP'
    str3 = ','
    str4 = 'CVP'

    str_list = [] #cria uma lista de comandos

    f = open(funcs.ficheiro, 'r', encoding='utf8')    #abre um ficheiro e devolve o conteudo

    while 1:
        line = f.readline()
        if not line:
            break
        #print (line) # do something

        ponto = line.split('\t', line.count('\t'))
        numero = ponto[0]
        coord_x = ponto[1]
        coord_y = ponto[2]
        if line.count('\t') > 2:
            obs = ponto[3]

        #numero, coord_x, coord_y = line.split('\t', 2)

        bloco = ''

        if str1 in numero:
            layer = 'rede_aerea'
        if str2 in numero:
            layer = 'rede_edp'
        if str4 in numero:
            layer = 'rede_subsolo'

        #if 'oste' in numero:
        if str4 in numero:
            bloco = 'cvp_existente'
        else:
            bloco = 'poste_existente'

        if str3 in coord_x:
            coord_x = coord_x.replace(',', '.')
        if str3 in coord_y:
            coord_y = coord_y.replace(',', '.')
        if '\n' in coord_y:
            coord_y = coord_y.replace('\n', '')

        texto1 = 'clayer\n' + layer + '\n_.insert\n'
        texto2 = bloco + '\n' + coord_x + ',' + coord_y + ',0\n1\n1\n0\n'

        texto = (texto1 + texto2)
        str_list.append(texto) #apende mais um comando à lista de comandos

    texto = 'clayer\n0\n'
    str_list.append(texto) #apende mais um comando à lista de comandos

    f.close()

    cad = ''.join(str_list) #junta os comandos todos

    f2 = open(funcs.pasta + 'blocos_rede_nos.scr', 'w')   #cria o ficheiro com comandos de autocad
    f2.write(cad)
    f2.close()

    mensagem = ('\nConversao ' + opcao + ' terminada com sucesso.\nVer ficheiro .scr na mesma pasta.')
    #funcs.escrever(mensagem)
    tkinter.messagebox.showinfo('Sucesso', mensagem)

def edif_pt():


    #-----recolher informacao
    
    tkinter.messagebox.showinfo('Abertura de ficheiro', 'Seleccionar ficheiro resumo com\ndados de survey.')

    opcao = 'edificios PT'
    #mensagem='\nEscolha: '+opcao
    #funcs.escrever(mensagem)

    funcs.t_fich="excel files"
    funcs.e_fich="*.xls"
   
    funcs.abrir()
    if funcs.ficheiro == '':
        tkinter.messagebox.showinfo('Erro', 'A abertura do ficheiro falhou')
        return 0

    str_list = [] #cria uma lista de comandos

    #f = open(funcs.ficheiro, 'r', encoding='utf8')    #abre um ficheiro e devolve o conteudo
    fich1 = funcs.ficheiro
    pontos = fich1.count('.')
    ext = fich1.split('.', pontos)
    extensao = (ext[pontos])
    if extensao == 'xls':
        funcs.open_xls_as_xlsx(fich1)
        wb1 = load_workbook(funcs.fich_novo)
    else:
        wb1 = load_workbook(funcs.ficheiro)
    ws1 = wb1.active
    #lin=1
    #col=1
    #x=ws1.cell(row=lin, column=col).value
    #print(x)

    array_resumo = []  #----------- inicio Resumo

    lin=6
    col_edif=6
    edif = ws1.cell(row=lin, column=col_edif).value
    
    while edif != None:   

        tipo_ed = ws1.cell(row=lin, column=col_edif+1).value
        pisos = ws1.cell(row=lin, column=col_edif+2).value
        ua_hab = ws1.cell(row=lin, column=col_edif+4).value
        ua_com = ws1.cell(row=lin, column=col_edif+5).value
        ua_esc = ws1.cell(row=lin, column=col_edif+6).value
        ua_ind = ws1.cell(row=lin, column=col_edif+7).value
        ua_esc = int(ua_esc) + int(ua_ind)
        coord_x = ws1.cell(row=lin, column=col_edif+8).value
        coord_y = ws1.cell(row=lin, column=col_edif+9).value
        tup1 = (edif, tipo_ed, pisos, ua_hab, ua_com, ua_esc, coord_x, coord_y)
        #print(tup1)
        array_resumo.append(tup1)
        lin = lin + 1
        edif = ws1.cell(row=lin, column=col_edif).value

    #print (array_resumo[2]) corresponde a terceira posicao do array-> 0,1,2
    #os.remove(funcs.fich_novo) 

    array_pd = [] #----------- inicio PDs
    
    tkinter.messagebox.showinfo('Abertura de ficheiro', 'Seleccionar ficheiro PDs com\ndados de survey.')

    funcs.t_fich="excel files"
    funcs.e_fich="*.xls"

    funcs.abrir()
    if funcs.ficheiro == '':
        tkinter.messagebox.showinfo('Erro', 'A abertura do ficheiro falhou')
        return 0
                                
    #f = open(funcs.ficheiro, 'r', encoding='utf8')    #abre um ficheiro e devolve o conteudo
    fich1 = funcs.ficheiro
    pontos = fich1.count('.')
    ext = fich1.split('.', pontos)
    extensao = (ext[pontos])
    if extensao == 'xls':
        funcs.open_xls_as_xlsx(fich1)
        wb1 = load_workbook(funcs.fich_novo)
    else:
        wb1 = load_workbook(funcs.ficheiro)
    ws1 = wb1.active
    #lin=1
    #col=1
    #x=ws1.cell(row=lin, column=col).value
    #print(x)

    lin = 4
    col_edif = 2
    edif = ws1.cell(row=lin, column=col_edif).value

    while edif != None:

        pd = ws1.cell(row=lin, column=col_edif+4).value
        tipo_pd = ws1.cell(row=lin, column=col_edif+5).value
        tup1 = (edif, pd, tipo_pd)
        #print(tup1)        
        array_pd.append(tup1)
        lin = lin + 1
        edif = ws1.cell(row=lin, column=col_edif).value

    os.remove(funcs.fich_novo)

    #-----fim recolher informacao


    #-----blocos scr

    texto_blocos = []

    tamanho = len(array_resumo)
    i = 0

    while i < tamanho:

        edif, tipo_ed, pisos, ua_hab, ua_com, ua_esc, coord_x, coord_y = array_resumo[i]#.split(',')

        #tipo bloco
        bloco = 'CE'
        if 'Serv' in tipo_ed:
            tipo = 'ME_SERV'
            bloco = 'CEE'
        if 'Prod' in tipo_ed:
            tipo = 'ME_PROD'
            bloco = 'CEE'
        if 'Mult' in tipo_ed:
            tipo = 'MULTI EMP'
            bloco = 'CEE'
        if 'Lote' in tipo_ed:
            tipo = 'LOTE VAZIO'
            bloco = 'CEE'
        #fim tipo bloco

        #procurar pd
        p = 0
        tamanho_pd = len(array_pd)
        num_pd = 'xxxx'
        tipo_pd = 'xxxx'
        while p < tamanho_pd:
            edif_pd, pd, tipo_p = array_pd[p]
            if edif_pd == edif:
                num_pd = pd
                tipo_pd = tipo_p
            p = p + 1
        #fim procurar pd

        if bloco == 'CE':

            x1 = str('_.insert\n' + bloco + '\n')
            x2 = str(coord_x) + ',' + str(coord_y) + ',0\n'
            x3 = '1\n1\n0\n'
            ua_com = ua_com + ua_esc #total uas nao-residenciais
            x4 = str(edif) + '\n' + str(pisos) + '\n' + str(ua_hab) + '\n' + str(ua_com) + '\n'
            x5 = str(num_pd) + '\n' + str(tipo_pd) + '\n'
            texto_temp = x1 + x2 + x3 + x4 + x5
            texto_blocos.append(texto_temp)

        if bloco == 'CEE':

            x1 = str('_.insert\n' + bloco + '\n')
            x2 = str(coord_x) + ',' + str(coord_y) + ',0\n'
            x3 = '1\n1\n0\n'
            x4 = str(edif) + '\n' + str(ua_hab) + '\n' + str(ua_com) + '\n' + str(ua_esc) + '\n'
            x5 = str(tipo) + '\n' + str(num_pd) + '\n' + str(tipo_pd) + '\n'
            texto_temp = x1 + x2 + x3 + x4 + x5
            texto_blocos.append(texto_temp)
            
        i = i + 1          

    scr = ''.join(texto_blocos)

    f = open(funcs.pasta + 'blocos.scr', 'w')   #cria o ficheiro scr na mesma pasta
    f.write(scr)
    f.close()
    array_resumo = [] #limpa tudo
    array_pd = []
    texto_temp = ''
    texto_blocos = []
    tkinter.messagebox.showinfo('Sucesso!', 'Terminado')