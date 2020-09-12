import tkinter as tk
from tkinter.filedialog import FileDialog
import tkinter.messagebox
import ntplib
from time import ctime
import xlrd
import pyproj
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
import os
import sys
sep = os.sep
coord_list = []

def abrir():

    global ficheiro
    global pasta
    global nome_fich
    global t_fich
    global e_fich

    d_fich="%UserProfile%\Desktop" + sep
    #pasta_defeito = os.system("%UserProfile%\Desktop" + sep if os.name == 'nt' else "$HOME/Desktop")
    ficheiro = tk.filedialog.askopenfilename(title="Escolher ficheiro.", filetypes = ((t_fich , e_fich),("all files","*.*")), initialdir=d_fich)
    pasta = os.path.dirname(os.path.abspath(ficheiro)) + sep

    if ficheiro == '':
        xpto = 0
    else:
        nome_fich = os.path.basename(ficheiro)
        nome_fich, extensao = nome_fich.split(".", 1)

    return ficheiro

def settings():

    global caminho

    caminho = os.path.dirname(os.path.abspath("__file__")) + sep
    try:
        with open(caminho + 'settings.ini') as file:
            f = open(caminho + 'settings.ini', 'r')
            f.close()
    except IOError:
        f = open(caminho + 'settings.ini', 'w')
        actual = "29,N,Portugal,PT-TM06"
        f.write(actual)
        f.close()

    return caminho

def escrever(mensagem):
    print(mensagem)

    return mensagem

def ver_datum():

    global datum1

    c = open(caminho + 'settings.ini', 'r')
    line = c.readline()
    zona_texto = line.split(',')
    datum1 = zona_texto[3]
    c.close()

    return datum1

def seg():

    try:
        c = ntplib.NTPClient()
        response = c.request('pool.ntp.org')
        data_completa = ctime(response.tx_time)
        data_completa = data_completa.replace("  ", " ") #dia 01 eh representado como espaco_extra+1

        semana, mes, dia, hora, ano = data_completa.split(" ", 4)

        if int(ano) > 2018:
            sys.exit(0)

    except:
        sys.exit(0)

        #tkinter.messagebox.showinfo('Erro', 'Este programa possui livrarias que necessitam\nde ligação à rede para funcionar.')
        #sys.exit(0)

#def array_coord(resultado[0] , resultado[1]):
    #tup1=(resultado[0],resultado[1])
    #array_coord.append(tup1)

def open_xls_as_xlsx(filename):

    global fich_novo
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()
    
    for row in range(1, nrows+1):
        for col in range(1, ncols+1):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row-1, col-1)
    book1.save(filename = pasta + 'temp.xlsx')
    fich_novo = pasta + 'temp.xlsx'
    
    return fich_novo

def conva2b(texto_conv,tipo_conv):

    wgs84 = pyproj.Proj('+init=EPSG:4326') # LatLon with WGS84 datum used by GPS units and Google Earth
    pttm06 = pyproj.Proj('+proj=tmerc +lat_0=39.66825833333333 +lon_0=-8.133108333333334 +k=1 +x_0=0 +y_0=0 +ellps=GRS80 +units=m +no_defs')#PT-TM06/ETRS89 EPSG:3763
    datum73IPCC = pyproj.Proj('+proj=tmerc +lat_0=39.66666666666666 +lon_0=-8.131906111111112 +k=1 +x_0=180.598 +y_0=-86.98999999999999 +ellps=intl +units=m +no_defs')#Datum 73 Hayford Gauss IPCC "ESRI:102161"
    datum73IGeoE = pyproj.Proj('+proj=tmerc +lat_0=39.66666666666666 +lon_0=-8.131906111111112 +k=1 +x_0=200180.598 +y_0=299913.01 +ellps=intl +units=m +no_defs')#Datum_73_Hayford_Gauss_IGeoE "ESRI:102160"
    datumOSGB36 = pyproj.Proj('+proj=tmerc +lat_0=49 +lon_0=-2 +k=0.9996012717 +x_0=400000 +y_0=-100000 +ellps=airy +datum=OSGB36 +units=m +no_defs')#OSGB-1936-EPSG:27700
    datumLambert93 = pyproj.Proj('+proj=lcc +lat_1=49 +lat_2=44 +lat_0=46.5 +lon_0=3 +x_0=700000 +y_0=6600000 +ellps=GRS80 +towgs84=0,0,0,0,0,0,0 +units=m +no_defs')#Lambert-93-EPSG:2154

    global strcoord

    ver_datum()
    if "TM06" in datum1:
        sist1 = pttm06
    if "IPCC" in datum1:
        sist1 = datum73IPCC
    if "IGeoE" in datum1:
        sist1 = datum73IGeoE
    if "OSGB36" in datum1:
        sist1 = datumOSGB36
    if "Lambert93" in datum1:
        sist1 = datumLambert93
    sist2 = wgs84

    xi,yi = texto_conv.split(",", 1)
    
    if tipo_conv == 1:
        strcoord = pyproj.transform(sist1, sist2, xi, yi)

    if tipo_conv == 2:
        strcoord = pyproj.transform(sist2, sist1, xi, yi)

    return strcoord